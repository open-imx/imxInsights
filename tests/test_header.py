import tempfile
from imxInsights import ImxMultiRepo, ImxContainer
from openpyxl import load_workbook



def test_specs_on_report_v1200(
    imx_v1200_zip_instance: ImxContainer,
    imx_v1200_dir_instance: ImxContainer,
    imx_v1200_specs_csv: str
):
    multi_repo = ImxMultiRepo([
        imx_v1200_zip_instance,
        imx_v1200_dir_instance
    ])

    compare = multi_repo.compare(imx_v1200_zip_instance.container_id, imx_v1200_dir_instance.container_id)

    spec_csv_file = imx_v1200_specs_csv

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as file_path:
        compare.to_excel(file_path, header_file=spec_csv_file)

        wb = load_workbook(filename=file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            if sheet_name in ['info', 'meta-overview']:
                continue

            sheet = wb[sheet_name]
            tag = sheet.cell(row=1, column=1).value
            assert tag == 'tagname', "first cell should have tag name"

            max_col = 0
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=2, column=col).value is not None:
                    max_col = col

            for col in range(2, max_col + 1):
                row1_value = sheet.cell(row=1, column=col).value
                row2_value = sheet.cell(row=2, column=col).value
                if row1_value in [
                    'N_ITER', 'Variable', 'PowerConnectionCableFunction', 'RelayConnectionCableFunction', 'AtpType',
                    'D_STATIC', 'NC_CDDIFF',
                    'Q_DIFF', 'Q_FRONT', 'V_DIFF', 'V_STATIC', 'A_NVP12', 'A_NVP23', 'L_NVKRINT', 'M_NVKRINT',
                    'M_NVKVINT', 'NID_C', 'Q_NVKVINTSET', 'V_NVKVINT', 'M_LEVELTEXTDISPLAY', 'M_MODETEXTDISPLAY'
                ]:
                    continue
                elif row1_value:
                    assert row2_value.endswith(row1_value), "Value row 2 should contain value in row 1"




