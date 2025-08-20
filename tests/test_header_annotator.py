import tempfile

from imxInsights import ImxMultiRepo, ImxContainer, ImxSingleFile
from openpyxl import load_workbook

from imxInsights.utils.headerAnnotator import HeaderSpec


def test_specs_on_report_v124(
        imx_v124_project_instance: ImxSingleFile,
        imx_v124_specs_csv: str
):
    multi_repo = ImxMultiRepo([
        imx_v124_project_instance.initial_situation,
        imx_v124_project_instance.new_situation
    ])

    compare = multi_repo.compare(
        imx_v124_project_instance.initial_situation.container_id,
        imx_v124_project_instance.new_situation.container_id
    )

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as file_path:
        file_path = "tester_v124.xlsx"
        compare.to_excel(file_path, header_spec=HeaderSpec(imx_v124_specs_csv))
        _check_header_for_specs(file_path, ['AtpType'])


def test_specs_on_report_v1200(
    imx_v1200_zip_instance: ImxContainer,
    imx_v1200_dir_instance: ImxContainer,
    imx_v1200_specs_csv: str
):
    multi_repo = ImxMultiRepo([
        imx_v1200_zip_instance,
        imx_v1200_dir_instance
    ])

    compare = multi_repo.compare(imx_v1200_zip_instance.container_id, imx_v1200_dir_instance.container_id)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as file_path:
        file_path = "tester_v1200.xlsx"
        compare.to_excel(file_path, header_spec=HeaderSpec(imx_v1200_specs_csv))

        # TODO: seems not to work for these columns.... not sure why
        _check_header_for_specs(file_path, [
            'N_ITER', 'Variable', 'PowerConnectionCableFunction', 'RelayConnectionCableFunction', 'AtpType',
            'D_STATIC', 'NC_CDDIFF', 'Q_DIFF', 'Q_FRONT', 'V_DIFF', 'V_STATIC', 'A_NVP12', 'A_NVP23', 'NID_C',
            'L_NVKRINT', 'M_NVKRINT', 'M_NVKVINT', 'Q_NVKVINTSET', 'V_NVKVINT', 'M_LEVELTEXTDISPLAY', 'M_MODETEXTDISPLAY'
        ])


def _check_header_for_specs(excel_file, columns_not_to_check):
    wb = load_workbook(filename=excel_file, data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name in ['info', 'meta-overview']:
            continue

        sheet = wb[sheet_name]
        tag = sheet.cell(row=1, column=1).value
        assert tag == 'tagname', "first cell should have tag name"

        max_col = 0
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=2, column=col).value is not None:
                max_col = col

        for col in range(2, max_col + 1):
            row1_value = sheet.cell(row=1, column=col).value
            row2_value = sheet.cell(row=2, column=col).value
            if row1_value in columns_not_to_check:
                continue
            elif row1_value:
                assert row2_value.endswith(row1_value), (
                    f"[{sheet_name}] Column {col}: Row 2 value '{row2_value}' does not end with Row 1 value '{row1_value}'"
                )