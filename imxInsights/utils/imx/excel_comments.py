from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from imxInsights.utils.report_helpers import REVIEW_STYLES

ExcelCell = Cell | MergedCell


def get_cell_background_color(cell: ExcelCell) -> str | None:
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == "rgb":
        return cell.fill.fgColor.rgb[-6:]
    return None


def get_column_indices(ws: Worksheet, header_row: int) -> dict[str, int | None]:
    column_map: dict[str, int | None] = {
        "@puic": None,
        "path": None,
        "status": None,
        "geometry_status": None,
    }
    for col in range(1, ws.max_column + 1):
        value = ws[f"{get_column_letter(col)}{header_row}"].value
        if value in column_map:
            column_map[value] = col
    return column_map


def get_cell_context(
    ws: Worksheet, row_idx: int, columns: dict[str, int | None]
) -> dict[str, Any]:
    return {
        "Puic": ws.cell(row=row_idx, column=columns["@puic"]).value
        if columns["@puic"]
        else None,
        "ObjectPath": ws.cell(row=row_idx, column=columns["path"]).value
        if columns["path"]
        else None,
        "ChangeStatus": ws.cell(row=row_idx, column=columns["status"]).value
        if columns["status"]
        else None,
        "GeometryStatus": ws.cell(row=row_idx, column=columns["geometry_status"]).value
        if columns["geometry_status"]
        else None,
    }


def handle_header_comment(
    cell: ExcelCell,
    ws: Worksheet,
    header_value: Any,
    columns: dict[str, int | None],
    sheetname: str,
    header_row: int,
) -> list[dict[str, Any]]:
    comments = []
    if cell.fill.fgColor and cell.fill.fgColor.rgb != "FFFFFFFF":
        for row_idx in range(header_row + 1, ws.max_row + 1):
            if not cell.column:
                continue
            target_cell = ws.cell(row=row_idx, column=cell.column)
            if (
                target_cell.fill.fgColor
                and target_cell.fill.fgColor.rgb != "FFFFFFFF"
                and target_cell.value
            ):
                context = get_cell_context(ws, row_idx, columns)

                if get_cell_background_color(target_cell) != "000000":
                    comments.append(
                        {
                            "Sheet": sheetname,
                            "Puic": context["Puic"],
                            "Header": header_value,
                            "Value": str(target_cell.value),
                            "Comment": cell.comment.text if cell.comment else "",
                            "CellAddress": target_cell.coordinate,
                            "Color": get_cell_background_color(target_cell),
                            "ObjectPath": context["ObjectPath"],
                            "ChangeStatus": context["ChangeStatus"],
                            "GeometryStatus": context["GeometryStatus"],
                            "CommentSheetName": sheetname,
                            "CommentRow": cell.row,
                            "CommentColumn": cell.column,
                        }
                    )
    return comments


def handle_data_comment(
    cell: ExcelCell,
    ws: Worksheet,
    header_value: Any,
    columns: dict[str, int | None],
    sheetname: str,
) -> list[dict[str, Any]]:
    if cell.row:
        context = get_cell_context(ws, cell.row, columns)
        return [
            {
                "Sheet": sheetname,
                "Puic": context["Puic"],
                "Header": header_value,
                "Value": str(cell.value),
                "Comment": cell.comment.text if cell.comment else "",
                "CellAddress": cell.coordinate,
                "Color": get_cell_background_color(cell),
                "ObjectPath": context["ObjectPath"],
                "ChangeStatus": context["ChangeStatus"],
                "GeometryStatus": context["GeometryStatus"],
                "CommentSheetName": sheetname,
                "CommentRow": cell.row,
                "CommentColumn": cell.column,
            }
        ]

    return []


def write_comments_sheet(wb: Workbook, comments: list[dict[str, Any]]) -> None:
    if "Comments" in wb.sheetnames:
        del wb["Comments"]
    ws = wb.create_sheet("Comments")
    ws.append(
        [
            "ObjectPath",
            "Puic",
            "ChangeStatus",
            "GeometryStatus",
            "Link",
            "ImxPath",
            "Value",
            "Comment",
            "Color",
            "CommentSheetName",
            "CommentRow",
            "CommentColumn",
        ]
    )
    color_to_status = {v: k for k, v in REVIEW_STYLES.items()}

    for comment in comments:
        link_text = color_to_status.get(comment["Color"], "link")
        link = f'=HYPERLINK("#\'{comment["Sheet"]}\'!{comment["CellAddress"]}", "{link_text}")'
        row = [
            comment["ObjectPath"],
            comment["Puic"],
            comment["ChangeStatus"],
            comment["GeometryStatus"],
            link,
            comment["Header"],
            comment["Value"],
            comment["Comment"],
            comment["Color"],
            comment["CommentSheetName"],
            comment["CommentRow"],
            comment["CommentColumn"],
        ]
        ws.append(row)
        if comment["Color"]:
            ws[f"E{ws.max_row}"].fill = PatternFill(
                start_color=comment["Color"],
                end_color=comment["Color"],
                fill_type="solid",
            )
    ws.freeze_panes = "A2"
    for col in range(1, ws.max_column + 1):
        length = (
            max(
                len(str(cell.value) or "")
                for cell in list(ws.iter_cols(min_col=col, max_col=col))[0]
            )
            + 4
        )
        ws.column_dimensions[get_column_letter(col)].width = length


def extract_comments_to_new_sheet(
    file_path: str, output_path: str | None = None, header_row: int = 1
) -> None:
    original_wb = load_workbook(file_path, data_only=True)
    all_comments = []

    for sheetname in original_wb.sheetnames:
        ws = original_wb[sheetname]
        columns = get_column_indices(ws, header_row)

        for row in ws.iter_rows():
            for cell in row:
                if cell.comment and cell.column:
                    header_cell = ws[f"{get_column_letter(cell.column)}{header_row}"]
                    header_value = header_cell.value if header_cell else None
                    if cell.row == header_row:
                        all_comments.extend(
                            handle_header_comment(
                                cell, ws, header_value, columns, sheetname, header_row
                            )
                        )
                    else:
                        all_comments.extend(
                            handle_data_comment(
                                cell, ws, header_value, columns, sheetname
                            )
                        )

    if output_path:
        wb = Workbook()
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)
    else:
        wb = original_wb
        if "Comments" in wb.sheetnames:
            del wb["Comments"]

    write_comments_sheet(wb, all_comments)

    save_path = output_path or file_path
    wb.save(save_path)
    print(f"Comments extracted successfully to '{save_path}' in sheet 'Comments'.")


# Example usage
# extract_comments_to_new_sheet("path to excel file", "optional_path_to_comment_file_else_add_sheet_to_input")
