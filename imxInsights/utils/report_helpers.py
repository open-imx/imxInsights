import importlib.metadata
import os
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, PatternFill
from pandas.io.formats.style import Styler
from xlsxwriter.worksheet import Worksheet  # type: ignore

INVALID_SHEET_CHARS = set(r"[]:*?/\\")


def sanitize_sheet_name(name: str) -> str:
    r"""
    Sanitize a string so it can be used as a valid Excel sheet name.

    Rules applied:
    - Remove invalid Excel characters (\ , /, ?, *, [, ], :).
    - Strip trailing apostrophes (Excel doesn’t allow a sheet name to end with `'`).
    - If the result is empty, fall back to "Sheet".

    Args:
        name (str): Proposed sheet name.

    Returns:
        str: A sanitized sheet name safe for Excel.
    """
    name = "".join(ch for ch in name if ch not in INVALID_SHEET_CHARS)
    name = name.rstrip("'") or "Sheet"
    return name


def shorten_sheet_name(sheet_name: str) -> str:
    """
    Ensure an Excel sheet name is both valid and within the 31-character limit.

    Workflow:
    - Sanitize the name (remove invalid characters, trailing apostrophes).
    - If the name length is <= 31, return it unchanged.
    - If it exceeds 31 characters, shorten it by keeping:
        * the first 14 characters,
        * the last 14 characters,
        * separated by "..." in the middle.

    Args:
        sheet_name (str): Proposed sheet name.

    Returns:
        str: A valid sheet name that is guaranteed to fit Excel's 31-char limit.
    """
    sheet_name = sanitize_sheet_name(sheet_name)
    if len(sheet_name) <= 31:
        return sheet_name
    head = 14
    tail = 14
    return f"{sheet_name[:head]}...{sheet_name[-tail:]}"


def clean_diff_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean a diff-report DataFrame by removing diff markers from specific columns.

    This function standardizes fields that may contain diff artifacts
    (e.g., `+`, `-`, `++`, `--`) introduced during change comparisons.

    Cleaning rules:
    - For columns `@puic`, `tag`, and `path`:
      * Convert values to strings.
      * Strip leading '+' or '-' characters.
    - For columns `parent` and `children`:
      * Replace `"++"` and `"--"` markers with empty strings.

    Args:
        df (pd.DataFrame): Input DataFrame containing diff results.

    Returns:
        pd.DataFrame: A cleaned copy of the DataFrame with markers removed.
    """
    df = df.copy()

    # Columns that need lstrip
    for col in ["@puic", "tag", "path"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.lstrip("+-")

    # Columns that need replacement
    replacements = {"++": "", "--": ""}
    for col in ["parent", "children"]:
        if col in df.columns:
            df[col] = df[col].replace(replacements)

    return df


def lower_and_index_duplicates(strings: list[str] | set[str]) -> list[str]:
    counts: dict[str, int] = {}
    result: set[str] = set()

    for s in (s.lower() for s in strings):
        counts[s] = counts.get(s, 0) + 1
        result.add(f"{s}{counts[s]}" if counts[s] > 1 else s)

    return sorted(result)


def upper_keys_with_index(original_dict: dict[str, Any]) -> dict[str, Any]:
    new_dict: dict[str, Any] = {}

    for key, value in original_dict.items():
        base_key = key.upper()
        new_key = base_key

        index = 1
        while new_key in new_dict:
            new_key = f"{base_key}_{index}"
            index += 1

        new_dict[new_key] = value

    return new_dict


def app_info_df(process_data: dict) -> pd.DataFrame:
    app_info = {
        "App Name": "ImxInsights",
        "Version": importlib.metadata.version("imxInsights"),
        "Developer": "OpenIMX",
        "Website": "https://open-imx.github.io/imxInsights/",
    }
    disclaimer = "This document is auto-generated. No guarantees are provided regarding the accuracy of the data."
    df = pd.DataFrame(list(app_info.items()), columns=["Attribute", "Value"])
    process_data_df = pd.DataFrame(
        list(process_data.items()), columns=["Attribute", "Value"]
    )

    disclaimer_df = pd.DataFrame({"Attribute": ["Disclaimer"], "Value": [disclaimer]})

    metadata_df = pd.concat(
        [
            df,
            pd.DataFrame([["", ""]]),
            process_data_df,
            pd.DataFrame([["", ""]]),
            disclaimer_df,
        ],
        ignore_index=True,
    )

    return metadata_df


def write_df_to_sheet(
    writer,
    sheet_name: str,
    df: pd.DataFrame | Styler,
    *,
    index: bool = False,
    header: bool = True,
    auto_filter: bool = True,
    grouped_columns: list[str] | None = None,
) -> Worksheet:
    """Write a DataFrame or Styler object to an Excel sheet."""
    df.to_excel(writer, sheet_name=sheet_name, index=index, header=header)
    worksheet = writer.sheets[sheet_name]
    worksheet.freeze_panes(1, 0)

    data = df.data if isinstance(df, Styler) else df  # type: ignore

    if auto_filter and not data.empty:
        num_cols = len(data.columns) - 1
        worksheet.autofilter(0, 0, 0, num_cols)

    if grouped_columns:
        for grouped_column in grouped_columns:
            worksheet.set_column(grouped_column, options={"level": 1, "hidden": True})

    worksheet.autofit()
    return worksheet


REVIEW_STYLES = {
    "OK": "80D462",
    "OK met opm": "66FF99",
    "NOK": "FF9999",
    "VRAAG": "F1F98F",
    "Bestaande fout": "FFCC66",
    "Aannemelijk": "E4DFEC",
    "TODO": "F2CEEF",
}


def add_review_styles_to_excel(file_name: str | Path) -> None:
    """
    Add predefined review styles as named styles to an existing Excel file.

    Args:
        file_name (str | Path): Path to the Excel file to modify.

    Returns:
        None
    """
    if isinstance(file_name, Path):
        file_name = f"{file_name}"

    wb = load_workbook(file_name)

    for name, color in REVIEW_STYLES.items():
        style = NamedStyle(name=name)
        style.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        if name not in wb.named_styles:
            wb.add_named_style(style)

    temp_file_name = tempfile.mktemp(suffix=".xlsx")
    wb.save(temp_file_name)

    with zipfile.ZipFile(temp_file_name, "r") as zip_in:
        with zipfile.ZipFile(file_name, "w") as zip_out:
            for item in zip_in.infolist():
                data = zip_in.read(item.filename)
                if item.filename == "xl/styles.xml":
                    tree = ET.fromstring(data)
                    ns = {
                        "ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                    }

                    idx_to_set = []
                    for idx, cell_style in enumerate(
                        tree.findall(".//ns:cellStyle", ns)
                    ):
                        if cell_style.attrib["name"] in REVIEW_STYLES.keys():
                            idx_to_set.append(idx)

                    for idx, cell_xf in enumerate(
                        tree.findall(".//ns:cellStyleXfs/ns:xf", ns)
                    ):
                        if idx in idx_to_set:
                            cell_xf.attrib["applyFont"] = "0"
                            cell_xf.attrib["applyBorder"] = "0"
                            cell_xf.attrib["applyAlignment"] = "0"
                            cell_xf.attrib["applyNumberFormat"] = "0"
                            cell_xf.attrib["applyFill"] = "1"

                    updated_data = ET.tostring(
                        tree, encoding="utf-8", xml_declaration=True
                    )
                    zip_out.writestr(item, updated_data)
                else:
                    zip_out.writestr(item, data)

    if os.path.exists(temp_file_name):
        os.remove(temp_file_name)


def add_nice_display(imx_object, props):
    # TODO: not sure, if we overwrite the ref, or add a column...
    add_column = True

    ref_lookup_map = {
        ref.lookup: ref.display for ref in imx_object.refs
    }  # Create a lookup map once

    result = {}
    for key, value in props.items():
        formatted_value = "\n".join(value.split(" "))
        result[key] = formatted_value

        if key.endswith("Ref"):
            if value in ref_lookup_map:
                result[f"{key}|display" if add_column else key] = ref_lookup_map[value]

        elif key.endswith("Refs"):
            ref_displays = [
                ref_lookup_map[item]
                for item in value.split(" ")
                if item in ref_lookup_map
            ]
            if ref_displays:
                result[f"{key}|display" if add_column else key] = "\n".join(
                    ref_displays
                )
    return result


def add_overview_df_to_diff_dict(
    diff_dict: dict[str, pd.DataFrame],
) -> dict[str, pd.DataFrame]:
    overview_df = pd.concat(list(diff_dict.values()), axis=0)
    columns_to_keep = [
        "@puic",
        "path",
        "tag",
        "ImxArea",
        "parent",
        "@name",
        "status",
        "geometry_status",
        "Location.GeographicLocation.@accuracy",
        "Location.GeographicLocation.@dataAcquisitionMethod",
        "Metadata.@isInService",
        "Metadata.@lifeCycleStatus",
        "Metadata.@source",
    ]
    existing_columns = [col for col in columns_to_keep if col in overview_df.columns]
    return {"meta-overview": overview_df[existing_columns]} | diff_dict


def unwrap_df(df: pd.DataFrame | pd.io.formats.style.Styler) -> pd.DataFrame:
    if isinstance(df, pd.DataFrame):
        return df
    return getattr(df, "data")


def set_sheet_color_by_change_status(df: pd.DataFrame | Styler, work_sheet: Worksheet):
    status_column = unwrap_df(df)["status"]
    valid_statuses = [
        "added",
        "changed",
        "unchanged",
        "type_change",
        "removed",
    ]
    status_values = status_column[status_column.isin(valid_statuses)]
    if status_values.eq("unchanged").all():
        work_sheet.set_tab_color("gray")


def apply_autofilter(
    worksheet: Worksheet, start_row: int, data_df: pd.DataFrame
) -> None:
    """
    Apply an autofilter to the header row of the data block.

    Args:
        worksheet (Worksheet): Target worksheet.
        start_row (int): Zero-based row where the data header (column names) is written.
        data_df (pd.DataFrame): The *data* part (not including metadata rows).
    """
    if data_df is None or data_df.empty:
        return
    # Columns are written starting at col 0; last data col index:
    last_col_idx = max(0, len(data_df.columns) - 1)
    worksheet.autofilter(start_row, 0, start_row, last_col_idx)


def autosize_columns(
    worksheet: Worksheet,
    full_df: pd.DataFrame,
    data_start_row: int,
    *,
    min_width: int = 15,
    header_min_width: int = 80,
    padding: int = 2,
) -> None:
    """
    Autosize columns based on the visible (non-metadata) cell contents and header text.

    Args:
        worksheet (Worksheet): Target worksheet.
        full_df (pd.DataFrame): Full DataFrame written to the sheet (metadata + data).
        data_start_row (int): First row index where data (not metadata) starts.
        min_width (int, optional): Minimum width per column (characters). Default 15.
        header_min_width (int, optional): Minimum width cap based on the header length. Default 80.
        padding (int, optional): Extra width padding (characters). Default 2.
    """
    if full_df is None or full_df.empty:
        return

    for col_idx, col_name in enumerate(full_df.columns):
        visible_values = full_df[col_name].iloc[data_start_row:]
        # Compute max content length among visible values
        if not visible_values.empty:
            max_content_len = visible_values.astype(str).map(len).max()
        else:
            max_content_len = 0

        # Ensure we show the full header name; use a reasonable cap for huge columns
        # We cap to max(header_min_width, header_len) to keep long headers readable.
        header_len = len(str(col_name))
        header_cap = max(header_min_width, header_len)

        target_len = max(max_content_len, header_len, min_width)
        final_width = min(target_len, header_cap) + padding
        worksheet.set_column(col_idx, col_idx, final_width)
